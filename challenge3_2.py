# !usr/bin/env python  
# -*- coding:utf-8 _*-  
""" 
@author:dandan.zheng 
@file: challenge3_2.py 
@time: 2018/03/28 
"""

from openpyxl import load_workbook # 可以用来载入已有数据表格
from openpyxl import Workbook # 可以用来处理新的数据表格
import datetime  # 可以用来处理时间相关的数据

# 使用 load_workbook.get_sheet_by_name() 载入 Excel 文件中的不同的表格
# 使用 load_workbook.remove() 可以移除 Excel 文件中的不同的表格
# Workbook().save('name.xlsx') 可以将对应的 Workbook 保存到本地

max_row = 486


def combine():
    # 该函数可以用来处理原数据文件：
    # 1. 合并表格并写入的 combine 表中
    # 2. 保存原数据文件
    wb = load_workbook('courses.xlsx')
    student_sheet = wb['students']
    time_sheet = wb['time']
    combine_sheet = wb.create_sheet('combine')
    heading = ['创建时间', '课程名称', '学习人数', '学习时间']
    combine_sheet.append(heading)
    for row in range(2, max_row):
        for col in range(1, 4):
            combine_sheet.cell(row=row, column=col, value=student_sheet.cell(row=row, column=col).value)
            if col == 3:
                for r in range(2, max_row):
                    if student_sheet.cell(row=row, column=2).value == time_sheet.cell(row=r, column=2).value:
                        combine_sheet.cell(row=row, column=col + 1, value=time_sheet.cell(row=r, column=3).value)
    wb.save('courses.xlsx')


def split():
    # 该函数可以用来分割文件：
    # 1. 读取 combine 表中的数据
    # 2. 将数据按时间分割
    # 3. 写入不同的数据表中
    # 导入工作目录
    wb = load_workbook('courses.xlsx')
    combine_sheet = wb['combine']

    # 获取第一列数
    years = []
    for row in range(2, max_row):
        create_time = combine_sheet.cell(row=row, column=1).value
        create_year = create_time.year
        # 文件目录
        if create_year not in years:
            years.append(create_year)

    for year in years:
        wb = Workbook()
        ws = wb.active
        ws.title = str(year)
        heading = ['创建时间', '课程名称', '学习人数', '学习时间']
        ws.append(heading)
        for row in range(2, max_row):
            result = []
            if combine_sheet.cell(row=row, column=1).value.year == year:
                for i in range(1, 5):
                    result.append(combine_sheet.cell(row=row, column=i).value)
                ws.append(result)
        wb.save(str(year) + '.xlsx')


# 执行
if __name__ == '__main__':
    combine()
    split()