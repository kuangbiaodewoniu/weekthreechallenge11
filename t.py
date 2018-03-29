# !usr/bin/env python  
# -*- coding:utf-8 _*-  
""" 
@author:dandan.zheng 
@file: t.py 
@time: 2018/03/28 
"""
from datetime import datetime
from openpyxl import Workbook, load_workbook

# 导入工作目录
wb = load_workbook('courses.xlsx')
combine_sheet = wb['combine']

# 获取第一列数
years = []
for row in range(2, 486):
    create_time = combine_sheet.cell(row=row, column=1).value
    create_year = str(create_time.year)
    # 文件目录
    if create_year not in years:
        years.append(create_year)

for year in years:
    wb = Workbook()
    ws = wb.active
    ws.title = year
    heading = ['创建时间', '课程名称', '学习人数', '学习时间']
    ws.append(heading)
    for row in range(2, 486):
        result = []
        if combine_sheet.cell(row=row, column=1).value.year == int(year):
            for i in range(1,5):
                result.append(combine_sheet.cell(row=row, column=i).value)
            ws.append(result)
    wb.save(year + '.xlsx')


